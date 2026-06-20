import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, Any

def generate_csv_bytes(df: pd.DataFrame) -> bytes:
    """
    Converts a pandas DataFrame to CSV bytes for download.
    """
    return df.to_csv(index=False).encode("utf-8")

def generate_excel_report(cleaned_df: pd.DataFrame, enriched_df: pd.DataFrame, quality_report: Dict[str, Any]) -> bytes:
    """
    Generates a professionally formatted multi-tab Excel report containing:
    - Tab 1: "Executive Summary" (Business KPIs, Sentiment, Categories, Quality Health)
    - Tab 2: "Cleaned Feedback Data" (The sanitized, validated dataset with quality flags)
    - Tab 3: "Enriched Feedback Data" (The fully analyzed and categorized dataset)
    
    Returns a bytes buffer representing the Excel workbook.
    """
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Create Excel sheets
        # 1. Executive Summary Sheet
        # 2. Cleaned Data
        # 3. Enriched Data
        
        # Write primary sheets first
        cleaned_df.to_excel(writer, sheet_name="Cleaned Feedback Data", index=False)
        enriched_df.to_excel(writer, sheet_name="Enriched Feedback Data", index=False)
        
        # Get workbook context to build the Executive Summary
        workbook = writer.book
        
        # Add Executive Summary as the first tab
        summary_sheet = workbook.create_sheet(title="Executive Summary", index=0)
        summary_sheet.views.sheetView[0].showGridLines = True
        
        # Styling parameters (indigo theme)
        title_font = Font(name="Calibri", size=16, bold=True, color="1E293B")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        section_font = Font(name="Calibri", size=12, bold=True, color="1E293B")
        bold_font = Font(name="Calibri", size=11, bold=True, color="000000")
        regular_font = Font(name="Calibri", size=11, color="000000")
        
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid") # Indigo
        section_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid") # Slate light
        
        thin_side = Side(border_style="thin", color="CBD5E1")
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        # 1. Sheet Title
        summary_sheet["A1"] = "Customer Feedback Intelligence Executive Report"
        summary_sheet["A1"].font = title_font
        summary_sheet.row_dimensions[1].height = 30
        
        # Helper function to write tables on summary sheet
        def write_section_header(sheet, start_row, title, col_span=3):
            sheet.cell(row=start_row, column=1, value=title).font = section_font
            sheet.cell(row=start_row, column=1).fill = section_fill
            sheet.row_dimensions[start_row].height = 20
            # Merge cell range for aesthetic
            sheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=col_span)
            
        def write_table_headers(sheet, start_row, headers):
            for i, h in enumerate(headers):
                cell = sheet.cell(row=start_row, column=i+1, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = border_all
            sheet.row_dimensions[start_row].height = 22

        # --- Section 1: Data Quality Audit Summary ---
        write_section_header(summary_sheet, start_row=3, title="1. DATA QUALITY AUDIT SUMMARY", col_span=3)
        write_table_headers(summary_sheet, start_row=4, headers=["Data Quality Metric", "Record Count", "Percentage"])
        
        quality_rows = [
            ("Total Uploaded Records", quality_report.get("total_rows", 0), "100.0%"),
            ("Missing Feedback Text (Filtered)", quality_report.get("missing_feedback_count", 0), f"{quality_report.get('missing_feedback_pct', 0.0):.2f}%"),
            ("Junk / Meaningless Text (Filtered)", quality_report.get("junk_count", 0), f"{quality_report.get('junk_pct', 0.0):.2f}%"),
            ("Missing Ratings (Flagged)", quality_report.get("missing_rating_count", 0), f"{quality_report.get('missing_rating_pct', 0.0):.2f}%"),
            ("Missing Timestamps (Flagged)", quality_report.get("missing_timestamp_count", 0), f"{quality_report.get('missing_timestamp_pct', 0.0):.2f}%"),
            ("Exact Duplicate Rows (Removed)", quality_report.get("exact_duplicates_count", 0), f"{quality_report.get('exact_duplicates_pct', 0.0):.2f}%"),
            ("Duplicate Feedback Text (Flagged)", quality_report.get("duplicate_feedback_text_count", 0), f"{quality_report.get('duplicate_feedback_text_pct', 0.0):.2f}%"),
            ("Invalid Ratings (Out of [1-5] Bounds)", quality_report.get("invalid_rating_count", 0), f"{quality_report.get('invalid_rating_pct', 0.0):.2f}%"),
            ("Emoji-only Feedback Messages", quality_report.get("emoji_only_count", 0), f"{quality_report.get('emoji_only_pct', 0.0):.2f}%"),
            ("Timestamp Formatting Inconsistencies", quality_report.get("inconsistent_timestamp_count", 0), f"{quality_report.get('inconsistent_timestamp_pct', 0.0):.2f}%"),
            ("Overall Data Quality Health Score", f"{quality_report.get('quality_health_score', 0.0)}/100", "")
        ]
        
        current_row = 5
        for metric, val, pct in quality_rows:
            c1 = summary_sheet.cell(row=current_row, column=1, value=metric)
            c2 = summary_sheet.cell(row=current_row, column=2, value=val)
            c3 = summary_sheet.cell(row=current_row, column=3, value=pct)
            
            c1.font = bold_font if metric == "Overall Data Quality Health Score" or metric == "Total Uploaded Records" else regular_font
            c2.font = bold_font if metric == "Overall Data Quality Health Score" or metric == "Total Uploaded Records" else regular_font
            c3.font = bold_font if metric == "Overall Data Quality Health Score" else regular_font
            
            c1.border = border_all
            c2.border = border_all
            c3.border = border_all
            
            c2.alignment = Alignment(horizontal="right")
            c3.alignment = Alignment(horizontal="right")
            
            summary_sheet.row_dimensions[current_row].height = 18
            current_row += 1
            
        # --- Section 2: AI Sentiment Distribution ---
        current_row += 2
        write_section_header(summary_sheet, start_row=current_row, title="2. CUSTOMER SENTIMENT ANALYSIS", col_span=3)
        current_row += 1
        write_table_headers(summary_sheet, start_row=current_row, headers=["Sentiment Class", "Cleaned Count", "Percentage Breakdown"])
        
        sent_counts = enriched_df["sentiment"].value_counts()
        total_enriched = len(enriched_df)
        
        current_row += 1
        for sent in ["Positive", "Negative", "Neutral"]:
            cnt = sent_counts.get(sent, 0)
            pct = (cnt / total_enriched * 100) if total_enriched > 0 else 0.0
            
            c1 = summary_sheet.cell(row=current_row, column=1, value=sent)
            c2 = summary_sheet.cell(row=current_row, column=2, value=int(cnt))
            c3 = summary_sheet.cell(row=current_row, column=3, value=f"{pct:.2f}%")
            
            c1.font = regular_font
            c2.font = regular_font
            c3.font = regular_font
            
            c1.border = border_all
            c2.border = border_all
            c3.border = border_all
            
            c2.alignment = Alignment(horizontal="right")
            c3.alignment = Alignment(horizontal="right")
            current_row += 1

        # --- Section 3: Feedback Category Distribution ---
        current_row += 2
        write_section_header(summary_sheet, start_row=current_row, title="3. FEEDBACK DOMAIN CATEGORIZATION", col_span=3)
        current_row += 1
        write_table_headers(summary_sheet, start_row=current_row, headers=["Category Domain", "Complaint Count", "Percentage Breakdown"])
        
        cat_counts = enriched_df["category"].value_counts()
        
        current_row += 1
        for cat in ["Billing", "App Bug", "Delivery", "Staff/Support", "Other"]:
            cnt = cat_counts.get(cat, 0)
            pct = (cnt / total_enriched * 100) if total_enriched > 0 else 0.0
            
            c1 = summary_sheet.cell(row=current_row, column=1, value=cat)
            c2 = summary_sheet.cell(row=current_row, column=2, value=int(cnt))
            c3 = summary_sheet.cell(row=current_row, column=3, value=f"{pct:.2f}%")
            
            c1.font = regular_font
            c2.font = regular_font
            c3.font = regular_font
            
            c1.border = border_all
            c2.border = border_all
            c3.border = border_all
            
            c2.alignment = Alignment(horizontal="right")
            c3.alignment = Alignment(horizontal="right")
            current_row += 1

        # --- Section 4: Hidden Signals (Sarcasm & Contradiction) ---
        current_row += 2
        write_section_header(summary_sheet, start_row=current_row, title="4. CRITICAL FEEDBACK ALERTS", col_span=3)
        current_row += 1
        write_table_headers(summary_sheet, start_row=current_row, headers=["Feedback Alert", "Flagged Count", "Percentage of Total"])
        
        sarcasm_cnt = enriched_df["sarcasm_flag"].sum()
        contra_cnt = enriched_df["contradiction_flag"].sum()
        
        alerts = [
            ("Sarcastic Customer Reviews", sarcasm_cnt, f"{(sarcasm_cnt / total_enriched * 100) if total_enriched > 0 else 0.0:.2f}%"),
            ("Contradictory Sentiment vs Rating Reviews", contra_cnt, f"{(contra_cnt / total_enriched * 100) if total_enriched > 0 else 0.0:.2f}%")
        ]
        
        current_row += 1
        for metric, cnt, pct in alerts:
            c1 = summary_sheet.cell(row=current_row, column=1, value=metric)
            c2 = summary_sheet.cell(row=current_row, column=2, value=int(cnt))
            c3 = summary_sheet.cell(row=current_row, column=3, value=pct)
            
            c1.font = regular_font
            c2.font = regular_font
            c3.font = regular_font
            
            c1.border = border_all
            c2.border = border_all
            c3.border = border_all
            
            c2.alignment = Alignment(horizontal="right")
            c3.alignment = Alignment(horizontal="right")
            current_row += 1

        # Auto-fit column widths for all sheets to make the report look clean
        for ws in workbook.worksheets:
            ws.views.sheetView[0].showGridLines = True
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    # Ignore cells that are merged or long text to prevent super wide columns
                    if cell.coordinate in ws.merged_cells:
                        continue
                    val_str = str(cell.value or '')
                    if len(val_str) > 50:
                        val_str = val_str[:47] + "..."
                    max_len = max(max_len, len(val_str))
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
    # Return Excel bytes
    return output.getvalue()
